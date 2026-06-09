from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from uuid import UUID
import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import get_db
from app.api.deps import require_seller_or_admin
from app.models.raffle import Raffle
from app.models.ticket import Ticket, TicketStatus
from app.models.payment import Payment, PaymentStatus

router = APIRouter(prefix="/reports", tags=["Reportes"])

# ── Styling constants ─────────────────────────────────────────────────────────
_INDIGO  = "4F46E5"
_VIOLET  = "7C3AED"
_LIGHT   = "EEEFFF"
_WHITE   = "FFFFFF"
_DARK    = "1E1B4B"
_MUTED   = "64748B"
_SUCCESS = "16A34A"
_SUCCESS_LIGHT = "F0FDF4"
_BORDER_COLOR  = "E2E8F0"

_HEADER_FONT  = Font(name="Calibri", size=11, bold=True, color=_WHITE)
_TITLE_FONT   = Font(name="Calibri", size=14, bold=True, color=_DARK)
_SUBTITLE_FONT= Font(name="Calibri", size=10, color=_MUTED)
_BODY_FONT    = Font(name="Calibri", size=10, color=_DARK)
_TOTAL_FONT   = Font(name="Calibri", size=10, bold=True, color=_DARK)

_thin = Side(style="thin", color=_BORDER_COLOR)
_CELL_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _col_fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color)


def _set_cols(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _header_row(ws, row: int, headers: list[str], color: str = _INDIGO) -> None:
    fill = _col_fill(color)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _HEADER_FONT
        c.fill = fill
        c.alignment = center
        c.border = _CELL_BORDER
    ws.row_dimensions[row].height = 26


def _data_cell(ws, row: int, col: int, value, align: str = "left", alt: bool = False) -> None:
    fill = _col_fill(_LIGHT) if alt else _col_fill(_WHITE)
    c = ws.cell(row=row, column=col, value=value)
    c.font = _BODY_FONT
    c.fill = fill
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border = _CELL_BORDER
    ws.row_dimensions[row].height = 20


# ── Buyers Excel ──────────────────────────────────────────────────────────────

@router.get("/raffles/{raffle_id}/buyers/csv")
def export_buyers(
    raffle_id: UUID,
    db: Session = Depends(get_db),
    _=Depends(require_seller_or_admin),
):
    raffle = db.query(Raffle).filter(Raffle.id == raffle_id).first()
    tickets = (
        db.query(Ticket)
        .filter(Ticket.raffle_id == raffle_id, Ticket.status == TicketStatus.paid)
        .order_by(Ticket.number)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Compradores"

    # ── Title block ──
    raffle_title = raffle.title if raffle else "Rifa"
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"Compradores — {raffle_title}"
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _col_fill(_LIGHT)
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    now = datetime.now(timezone.utc).strftime("%d/%m/%Y a las %H:%M UTC")
    c.value = f"Generado el {now}  ·  {len(tickets)} boletos vendidos"
    c.font = _SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    ws.row_dimensions[3].height = 8

    # ── Headers ──
    _header_row(ws, 4, ["Boleto #", "Nombre completo", "Teléfono", "Email", "Fecha de pago"])
    _set_cols(ws, [12, 32, 22, 30, 24])

    # ── Data rows ──
    for i, t in enumerate(tickets, 5):
        alt = i % 2 == 0
        paid_str = t.paid_at.strftime("%d/%m/%Y %H:%M") if t.paid_at else "—"
        _data_cell(ws, i, 1, f"#{t.number:04d}", align="center", alt=alt)
        _data_cell(ws, i, 2, t.buyer_name  or "—", alt=alt)
        _data_cell(ws, i, 3, t.buyer_phone or "—", alt=alt)
        _data_cell(ws, i, 4, t.buyer_email or "—", alt=alt)
        _data_cell(ws, i, 5, paid_str,               alt=alt)

    # ── Totals footer ──
    total_row = len(tickets) + 5
    if tickets:
        ws.row_dimensions[total_row].height = 8
        total_row += 1

    ws.merge_cells(f"A{total_row}:B{total_row}")
    c = ws[f"A{total_row}"]
    c.value = f"Total boletos vendidos: {len(tickets)}"
    c.font = _TOTAL_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = _col_fill(_LIGHT)

    if raffle and raffle.ticket_price:
        revenue = len(tickets) * raffle.ticket_price
        ws.merge_cells(f"C{total_row}:D{total_row}")
        c = ws[f"C{total_row}"]
        c.value = f"Recaudado: ${revenue:,.0f}"
        c.font = _TOTAL_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.fill = _col_fill(_LIGHT)

    ws.row_dimensions[total_row].height = 22
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=compradores-{raffle_id}.xlsx"},
    )


# ── Payments Excel ────────────────────────────────────────────────────────────

@router.get("/raffles/{raffle_id}/payments/csv")
def export_payments(
    raffle_id: UUID,
    db: Session = Depends(get_db),
    _=Depends(require_seller_or_admin),
):
    raffle = db.query(Raffle).filter(Raffle.id == raffle_id).first()
    payments = (
        db.query(Payment)
        .join(Ticket, Ticket.payment_id == Payment.id)
        .filter(Ticket.raffle_id == raffle_id)
        .distinct()
        .all()
    )

    _METHOD_LABELS = {
        "wompi": "Wompi", "mercadopago": "MercadoPago",
        "cash": "Efectivo", "transfer": "Transferencia",
    }
    _STATUS_LABELS = {
        "pending": "Pendiente", "confirmed": "Confirmado",
        "failed": "Fallido", "refunded": "Reembolsado",
    }
    _STATUS_COLORS = {
        "confirmed": _SUCCESS_LIGHT,
        "pending":   "FFFBEB",
        "failed":    "FEF2F2",
        "refunded":  "F0F9FF",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    raffle_title = raffle.title if raffle else "Rifa"
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"Pagos — {raffle_title}"
    c.font = _TITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = _col_fill(_LIGHT)
    ws.row_dimensions[1].height = 32

    confirmed = sum(1 for p in payments if p.status.value == "confirmed")
    total_amount = sum(p.amount for p in payments if p.status.value == "confirmed")

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    now = datetime.now(timezone.utc).strftime("%d/%m/%Y a las %H:%M UTC")
    c.value = f"Generado el {now}  ·  {len(payments)} pagos totales · {confirmed} confirmados"
    c.font = _SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 8

    _header_row(ws, 4, ["ID (primeros 8)", "Monto", "Método", "Estado", "Referencia", "Fecha", "Boletos"], color=_VIOLET)
    _set_cols(ws, [14, 14, 18, 16, 22, 22, 10])

    for i, p in enumerate(payments, 5):
        alt = i % 2 == 0
        status_val = p.status.value if hasattr(p.status, "value") else str(p.status)
        method_val = p.method.value if hasattr(p.method, "value") else str(p.method)
        row_fill = _col_fill(_STATUS_COLORS.get(status_val, _WHITE))

        ticket_count = sum(1 for t in p.tickets) if hasattr(p, "tickets") else "—"

        cells = [
            (str(p.id)[:8].upper(), "center"),
            (f"${p.amount:,.0f}", "right"),
            (_METHOD_LABELS.get(method_val, method_val), "center"),
            (_STATUS_LABELS.get(status_val, status_val), "center"),
            (p.gateway_reference or "—", "center"),
            (p.created_at.strftime("%d/%m/%Y %H:%M") if p.created_at else "—", "center"),
            (str(ticket_count), "center"),
        ]
        for col, (val, align) in enumerate(cells, 1):
            c = ws.cell(row=i, column=col, value=val)
            c.font = _BODY_FONT
            c.fill = row_fill
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.border = _CELL_BORDER
        ws.row_dimensions[i].height = 20

    # Footer
    total_row = len(payments) + 5
    if payments:
        ws.row_dimensions[total_row].height = 8
        total_row += 1

    ws.merge_cells(f"A{total_row}:C{total_row}")
    c = ws[f"A{total_row}"]
    c.value = f"Total confirmado: ${total_amount:,.0f}"
    c.font = _TOTAL_FONT
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.fill = _col_fill(_LIGHT)
    ws.row_dimensions[total_row].height = 22
    ws.freeze_panes = "A5"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pagos-{raffle_id}.xlsx"},
    )
